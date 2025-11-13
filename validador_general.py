# ===============================================
# PASO 1: SUBIDA Y VALIDACIÓN DEL ARCHIVO 1
# ===============================================

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment

# ================================================
# CONFIGURACIÓN INICIAL
# ================================================
st.set_page_config(
    page_title="Validador de Archivos",
    page_icon="📊",
    layout="wide"
)

# ================================================
# INICIALIZACIÓN DE ESTADOS
# ================================================
if "paso_actual" not in st.session_state:
    st.session_state.paso_actual = 0
if "nombre_colegio" not in st.session_state:
    st.session_state.nombre_colegio = ""
if "comparador_archivo_base" not in st.session_state:
    st.session_state.comparador_archivo_base = None
if "comparador_archivo_revisar" not in st.session_state:
    st.session_state.comparador_archivo_revisar = None
if "comparador_resultados" not in st.session_state:
    st.session_state.comparador_resultados = None
if "archivo1_df" not in st.session_state:
    st.session_state.archivo1_df = None
if "archivo2_df" not in st.session_state:
    st.session_state.archivo2_df = None
if "archivo2_1p3p_df" not in st.session_state:
    st.session_state.archivo2_1p3p_df = None
if "archivo2_4p5s_df" not in st.session_state:
    st.session_state.archivo2_4p5s_df = None
if "archivo1_bytes" not in st.session_state:
    st.session_state.archivo1_bytes = None
if "archivo2_bytes" not in st.session_state:
    st.session_state.archivo2_bytes = None
if "archivo1_fila_cabecera" not in st.session_state:
    st.session_state.archivo1_fila_cabecera = None
if "archivo2_1p3p_fila_cabecera" not in st.session_state:
    st.session_state.archivo2_1p3p_fila_cabecera = None
if "archivo2_4p5s_fila_cabecera" not in st.session_state:
    st.session_state.archivo2_4p5s_fila_cabecera = None
if "cursos_equivalentes" not in st.session_state:
    st.session_state.cursos_equivalentes = [
        "ADOBE ILLUSTRATOR", "ADOBE INDESIGN", "ADOBE PHOTOSHOP PROFICIENT",
        "COACHING PERSONAL Y VOCACIONAL", "DE LA IDEA AL EMPRENDIMIENTO",
        "DESARROLLO DE APLICACIONES MÓVILES", "DESARROLLO WEB", "DISEÑO WEB",
        "EDICIÓN DE AUDIO", "EDICIÓN DE VIDEO", "EXCEL EXPERT SPECIALIST",
        "EXCEL INTERMEDIATE SPECIALIST", "EXCEL PROFICIENT SPECIALIST",
        "FINANZAS PERSONALES", "GESTIÓN DE DATA CON MS EXCEL Y POWER BI",
        "GESTIÓN EMPRESARIAL", "HABILIDADES BLANDAS", "MARKETING DIGITAL",
        "MARKETING PERSONAL", "PRESENTACIONES DE IMPACTO",
        "PROGRAMACIÓN VISUAL KODU PLANET I", "PROGRAMACIÓN VISUAL KODU PLANET II",
        "PROGRAMACIÓN VISUAL KODU PLANET III", "ROBLOX FOR TEENS", "ROBÓTICA",
        "SCRATCH", "WORD EXPERT SPECIALIST", "WORD PROFICIENT SPECIALIST",
        "LEARNING FOR BEGINNERS 1", "LEARNING FOR BEGINNERS 2", "CODE FOR KIDS"
    ]

# ================================================
# CONSTANTES
# ================================================
COLUMNAS_ARCHIVO1 = [
    "Nro.", "Paterno", "Materno", "Nombres", "Nacimiento (DD/MM/YYYY)", "Sexo (M/F)",
    "Grado", "Sección", "Correo institucional", "Neurodiversidad (Sí/No)", "DNI"
]

COLUMNAS_ARCHIVO2 = [
    "Nro.", "Paterno", "Materno", "Nombres", "Curso", "Grado", "Sección", "Nota Vigesimal"
]

COLUMNAS_EVALUADOR = [
    "Nro.", "Paterno", "Materno", "Nombres", "Curso", "Grado", "Sección", 
    "Nota Vigesimal", "NOTAS VIGESIMALES 75%", "PROMEDIO", "OBSERVACIONES"
]

# Constantes de validación
SEXO_VALIDO = ["M", "F"]
SECCIONES_VALIDAS = ["A", "B", "C", "D", "E", "F", "G", "U", "UNICO", "UNICA", "ÚNICO", "ÚNICA", "Único", "Única"]
GRADOS_VALIDOS = ["1P", "2P", "3P", "4P", "5P", "6P", "1S", "2S", "3S", "4S", "5S"]
MAPEO_GRADOS = {
    "1": "1P", "2": "2P", "3": "3P", "4": "4P", "5": "5P", "6": "6P",
    "7": "1S", "8": "2S", "9": "3S", "10": "4S", "11": "5S"
}
MAPEO_SECCIONES = {
    "UNICO": "U",
    "UNICA": "U",
    "ÚNICO": "U", 
    "ÚNICA": "U", 
    "Único": "U", 
    "Única": "U"
}
LISTA_COLEGIOS = [
    "Colegio Ateneo la Molina",
    "Colegio Bárbara Dachille",
    "Colegio Bautista Filadelfia-Comas",
    "Colegio Cepeban",
    "Colegio Cervantes School",
    "Colegio Divino Maestro de Pro",
    "Colegio Divino Niño Jesús",
    "Colegio Don Bosco - San Luis ",
    "Colegio Dora Mayer",
    "Colegio El Carmelo",
    "Colegio Giordano Bruno",
    "Colegio Gracias Jesús",
    "Colegio Ingeniero Carlos Lisson Beingolea",
    "Colegio Innova America High School",
    "Colegio Joseph Novak",
    "Colegio Loris Malaguzzi",
    "Colegio Los Rosales de Santa Rosa",
    "Colegio Maestro Redentor - Huancayo ",
    "Colegio Mahatma Gandhi",
    "Colegio Makarenko",
    "Colegio María Inmaculada Concepción",
    "Colegio Mashal School",
    "Colegio Miguel Angel",
    "Colegio My Home And School",
    "Colegio Nuestra Señora Carmen de Palao",
    "Colegio Parroquial San José",
    "Colegio Redimer Jesús De Villa",
    "Colegio San Francisco -Balconcillo ",
    "Colegio San José de los Balnearios Del Sur",
    "Colegio San Martincito de Porres",
    "Colegio Santa Ana - Tacna",
    "Colegio Santa Angela Merici",
    "Colegio Santa María de Surco",
    "Colegio Signos de La Fe La Salle - Trujillo ",
    "Colegio Sor Querubina de San Pedro",
    "Colegio Andino Huancayo",
    "Colegio Andre Malraux",
    "Colegio Antares",
    "Colegio Divina Misericordia",
    "Colegio Ingeniería Huancayo",
    "Colegio Jesús el Nazareno",
    "Colegio Jesús María - San Martín de Porres",
    "Colegio Lima International School Of Tomorrow",
    "Colegio Lincoln del Triunfo",
    "Colegio Luz Casanova",
    "Colegio Magister",
    "Colegio María Montessori de Copacabana",
    "Colegio María Rafols",
    "Colegio Melvin Jones",
    "Colegio Nuestra Señora del Buen Consejo",
    "Colegio Patrocinio San José",
    "Colegio Peruano Japones La Victoria",
    "Colegio Play School Huaral",
    "Colegio San Antonio de Padua",
    "Colegio San Antonio María Claret",
    "Colegio San Charbel",
    "Colegio San Francisco de Borja",
    "Colegio San Germán",
    "Colegio San José Hermanos Maristas Callao",
    "Colegio San Mateo Anglicano",
    "Colegio Santa Ana - Lima",
    "Colegio Santa Angela",
    "Colegio Santa Anita",
    "Colegio Santa Rosa de Lima"
]

# ================================================
# FUNCIONES AUXILIARES
# ================================================
def detectar_cabecera_automatica(df: pd.DataFrame, columnas_objetivo: list):
    """Detecta automáticamente la fila de cabecera"""
    max_filas, max_cols = min(15, len(df)), min(15, len(df.columns))
    subset = df.iloc[:max_filas, :max_cols]
    columnas_objetivo_norm = [c.strip().lower() for c in columnas_objetivo]

    for idx in range(max_filas):
        fila = subset.iloc[idx].astype(str).str.strip().str.lower().tolist()
        if all(col in fila for col in columnas_objetivo_norm):
            return idx
    return None

def crear_identificador(df, col_paterno, col_materno, col_nombres):
    """Crea columna identificador"""
    return (
        df[col_nombres].astype(str).str.strip() + " " +
        df[col_paterno].astype(str).str.strip() + " " +
        df[col_materno].astype(str).str.strip()
    )

def normalizar_enie(texto):
    """
    Normaliza texto a mayúsculas preservando TODOS los acentos (tildes y Ñ)
    """
    if pd.isna(texto):
        return ""
    
    # Convertir a mayúsculas y limpiar espacios
    texto = str(texto).strip().upper()
    
    # Normalizar espacios múltiples
    texto = ' '.join(texto.split())
    
    return texto

def limpiar_filas_vacias(df, columnas_clave=None):
    """
    Args:
        df: DataFrame a limpiar
        columnas_clave: Lista de nombres de columnas para verificar (default: primeras 4)
    
    Returns:
        DataFrame limpio sin filas completamente vacías (evita sólo los Nro o N°)
    """
    if columnas_clave is None:
        # Usar las columnas 2, 3, 4
        columnas_clave = df.columns[1:4].tolist()
    
    # Contar registros originales
    total_original = len(df)
    
    # Filtrar: mantener solo filas con almenos una columna clave tenga datos
    df_limpio = df.dropna(subset=columnas_clave, how='all').copy()
    
    # Eliminar filas donde todas las columnas clave sean strings vacíos
    mask = df_limpio[columnas_clave].apply(
        lambda x: x.astype(str).str.strip().ne('')
    ).any(axis=1)
    df_limpio = df_limpio[mask].reset_index(drop=True)
    
    # Mostrar info si se eliminaron filas
    filas_eliminadas = total_original - len(df_limpio)
    if filas_eliminadas > 0:
        st.info(f"🧹 Se eliminaron {filas_eliminadas} filas vacías (quedaron {len(df_limpio)} registros)")
    
    return df_limpio

def homologar_dataframe(df):
    """
    Homologa un DataFrame completo:
    - Todas las columnas: Convierte a mayúsculas y quita espacios
    - Columnas PATERNO, MATERNO, NOMBRES: Además quita acentos y mantiene la Ñ
    """
    # Columnas especiales que requieren normalización de acentos
    columnas_nombres = ["PATERNO", "MATERNO", "NOMBRES"]
    filas_vacias = df[df[columnas_nombres].isnull().any(axis=1)]

    if not filas_vacias.empty:
        st.error("❌ Se detectaron campos vacíos en nombres o apellidos (Archivo 1 - Nómina)")
        st.dataframe(filas_vacias, use_container_width=True)
        st.stop()

    # Procesar todas las columnas
    for col in df.columns:
        if col.upper() in columnas_nombres:
            # Para columnas de nombres: usar función que preserva Ñ
            df[col] = df[col].apply(normalizar_enie)
            # Normalizar espacios múltiples
            df[col] = df[col].str.replace(r'\s+', ' ', regex=True).str.strip()
        else:
            # Solo mayúsculas y quitar espacios
            df[col] = (
                df[col].astype(str)
                .str.strip()
                .str.upper()
                .str.replace(r'\s+', ' ', regex=True)
                .str.strip()
            )
    
    return df

def convertir_numericas_a_entero(df, columnas=None):
    """
    Convierte valores numéricos flotantes a enteros (1.0 → "1")
    Funciona incluso en columnas mixtas (1.0, "2P", 3.0)
    
    Args:
        df: DataFrame a procesar
        columnas: Lista de columnas a convertir
    
    Returns:
        DataFrame con columnas convertidas
    """
    if columnas is None:
        columnas = df.select_dtypes(include=['float64', 'float32']).columns.tolist()
    
    for col in columnas:
        if col not in df.columns:
            continue
        
        # Procesar cada valor individualmente
        def convertir_valor(val):
            """Convierte un valor individual"""
            if pd.isna(val):
                return val
            
            # Convertir a string para inspeccionar
            val_str = str(val).strip()
            
            # Si ya tiene letras, dejarlo como está
            if any(c.isalpha() for c in val_str):
                return val_str
            
            # Si es numérico puro, intentar convertir
            try:
                val_num = float(val)
                # Si es un entero disfrazado de float (1.0, 2.0)
                if val_num % 1 == 0:
                    return str(int(val_num))
                else:
                    # Si tiene decimales reales (1.5), mantener como string
                    return val_str
            except (ValueError, TypeError):
                return val_str
        
        # Aplicar la conversión a toda la columna
        df[col] = df[col].apply(convertir_valor)
    
    return df

def validar_y_mapear_grados(df, col_grado="GRADO", tipo_validacion="todos"):
    """
    Valida y mapea los grados. Convierte números 1-11 a formato estándar (1P-6P, 1S-5S).
    Retorna DataFrame procesado y lista de errores.
    
    Args:
        df: DataFrame a validar
        col_grado: Nombre de la columna de grado
        tipo_validacion: Tipo de validación a aplicar:
            - "todos": Valida todos los grados (1P-6P, 1S-5S) - Para Archivo 1
            - "1p3p": Solo valida 1P, 2P, 3P - Para hoja 1P-3P del Archivo 2
            - "4p5s": Solo valida 4P-6P, 1S-5S - Para hoja 4P-5S del Archivo 2
    """
    errores = []
    df[col_grado] = df[col_grado].astype(str).str.strip().str.upper()
    
    # Definir mapeos según el tipo de validación
    if tipo_validacion == "1p3p":
        mapeo_grados = {
            "1": "1P", "2": "2P", "3": "3P"
        }
        grados_validos = ["1P", "2P", "3P"]
    elif tipo_validacion == "4p5s":
        mapeo_grados = {
            "4": "4P", "5": "5P", "6": "6P",
            "7": "1S", "8": "2S", "9": "3S", "10": "4S", "11": "5S"
        }
        grados_validos = ["4P", "5P", "6P", "1S", "2S", "3S", "4S", "5S"]
    else:  # "todos"
        mapeo_grados = MAPEO_GRADOS
        grados_validos = GRADOS_VALIDOS
    
    # Mapear números a grados
    df[col_grado] = df[col_grado].replace(mapeo_grados)
    
    # Validar grados
    grados_invalidos = df.loc[~df[col_grado].isin(grados_validos)]

    if len(grados_invalidos) > 0:
        for idx, row in grados_invalidos.iterrows():
            errores.append(f"Fila {idx + 2}: Grado inválido '{row[col_grado]}'")
    
    return df, errores

def inferir_sexo_por_nombre(nombre):
    """
    Infiere el sexo basándose en el nombre.
    Retorna 'M' o 'F' según terminaciones comunes en español.
    """
    if pd.isna(nombre) or str(nombre).strip() == "":
        return "M"  # Por defecto M si no hay nombre
    
    nombre = str(nombre).strip().upper()
    primer_nombre = nombre.split()[0] if nombre else ""
    
    # Terminaciones típicamente femeninas
    terminaciones_femeninas = ['A', 'IA', 'INA', 'ELA', 'ANA', 'LIA', 'RIA', 'TA', 'DA']
    # Nombres específicamente femeninos comunes
    nombres_femeninos = ['MARIA', 'CARMEN', 'ROSA', 'LUZ', 'SOL', 'MERCEDES', 'BEATRIZ', 'INES', 'ISABEL']
    
    if primer_nombre in nombres_femeninos:
        return "F"
    
    for term in terminaciones_femeninas:
        if primer_nombre.endswith(term):
            return "F"
    
    return "M"  # Por defecto masculino

def validar_sexo(df, col_sexo="SEXO (M/F)"):
    """
    Valida que el sexo sea M o F.
    Si está vacío, infiere el sexo según el nombre del alumno.
    Retorna lista de errores (ahora solo para casos que no se puedan resolver).
    """
    errores = []
    df[col_sexo] = df[col_sexo].astype(str).str.strip().str.upper()
    
    # Reemplazar valores vacíos o inválidos por inferencia basada en nombre
    mask_vacios_invalidos = ~df[col_sexo].isin(SEXO_VALIDO)
    
    if mask_vacios_invalidos.any():
        for idx in df[mask_vacios_invalidos].index:
            nombre = df.loc[idx, "NOMBRES"] if "NOMBRES" in df.columns else ""
            sexo_inferido = inferir_sexo_por_nombre(nombre)
            df.loc[idx, col_sexo] = sexo_inferido
            # Se registra como advertencia informativa (no error crítico)
            identificador = crear_identificador(df.loc[[idx]], "PATERNO", "MATERNO", "NOMBRES").iloc[0]
            errores.append(f"INFO - Fila {idx + 2}: Sexo vacío/inválido, se asignó '{sexo_inferido}' según nombre - {identificador}")
    return errores

def validar_secciones(df, col_seccion="SECCIÓN"):
    """
    Valida que las secciones sean válidas (A-G, U, UNICO, UNICA y estas dos últimas reemplazarlas por U).
    Retorna lista de errores.
    """
    errores = []
    df[col_seccion] = df[col_seccion].astype(str).str.strip().str.upper()
    
    # (UNICO/UNICA -> U)
    df[col_seccion] = df[col_seccion].replace(MAPEO_SECCIONES)

    secciones_invalidas = df.loc[~df[col_seccion].isin(SECCIONES_VALIDAS)]

    if len(secciones_invalidas) > 0:
        for idx, row in secciones_invalidas.iterrows():
            errores.append(f"Fila {idx + 2}: Sección inválida '{row[col_seccion]}' (debe ser A-G o U)")
    
    return errores

def validar_neurodiversidad(df, col_neuro="NEURODIVERSIDAD (SÍ/NO)"):
    """
    Valida que neurodiversidad sea Sí o No.
    Retorna lista de errores.
    """
    errores = []
    df[col_neuro] = df[col_neuro].astype(str).str.strip().str.upper()
    
    # Mapear variaciones comunes
    mapeo_neuro = {
        "SI": "SÍ", "S": "SÍ", "YES": "SÍ", "Y": "SÍ",
        "N": "NO", "NOT": "NO"
    }
    df[col_neuro] = df[col_neuro].replace(mapeo_neuro)
    
    neuros_invalidas = df.loc[~df[col_neuro].isin(["SÍ", "NO"])]

    if len(neuros_invalidas) > 0:
        for idx, row in neuros_invalidas.iterrows():
            identificador = crear_identificador(df.loc[[idx]], "PATERNO", "MATERNO", "NOMBRES").iloc[0]
            errores.append(f"Fila {idx + 2}: Neurodiversidad inválida '{row[col_neuro]}' - {identificador}")
    
    return errores

def validar_fecha_nacimiento(df, col_fecha="NACIMIENTO (DD/MM/YYYY)"):
    """
    Valida y formatea fechas al formato DD/MM/YYYY.
    Retorna lista de errores y modifica el DataFrame.
    """
    errores = []
    
    for idx, row in df.iterrows():
        fecha_original = str(row[col_fecha]).strip()
        identificador = crear_identificador(df.loc[[idx]], "PATERNO", "MATERNO", "NOMBRES").iloc[0]
        
        # Intentar parsear la fecha con diferentes formatos comunes
        fecha_parseada = pd.to_datetime(fecha_original, errors="coerce", dayfirst=True)
        
        if pd.isna(fecha_parseada):
            errores.append(f"Fila {idx + 2}: Fecha inválida '{fecha_original}' - {identificador}")
        else:
            # Formatear al formato deseado DD/MM/YYYY
            fecha_formateada = fecha_parseada.strftime("%d/%m/%Y")
            df.at[idx, col_fecha] = fecha_formateada
    
    return errores

def validar_dni(df, col_dni="DNI"):
    """
    Valida que el DNI tenga exactamente 8 dígitos.
    Retorna lista de errores.
    """
    errores = []
    df[col_dni] = df[col_dni].astype(str).str.strip()
    
    for idx, row in df.iterrows():
        dni = row[col_dni]
        identificador = crear_identificador(df.loc[[idx]], "PATERNO", "MATERNO", "NOMBRES").iloc[0]
        
        # Validar que sea número de 8 dígitos
        if not (dni.isdigit() and len(dni) == 8):
            errores.append(f"Fila {idx + 2}: DNI inválido '{dni}' (debe ser 8 dígitos) - {identificador}")
    
    return errores

def validar_correo(df, col_correo="CORREO INSTITUCIONAL"):
    """
    Valida formato básico de correo electrónico.
    Retorna lista de errores.
    """
    errores = []
    
    for idx, row in df.iterrows():
        correo = str(row[col_correo]).strip().lower()
        identificador = crear_identificador(df.loc[[idx]], "PATERNO", "MATERNO", "NOMBRES").iloc[0]
        
        # Validación básica: contiene @ y .
        if "@" not in correo or "." not in correo.split("@")[-1]:
            errores.append(f"Fila {idx + 2}: Correo inválido '{correo}' - {identificador}")
    
    return errores

def mostrar_stepper(paso_actual):
    """Muestra el indicador de progreso visual"""
    pasos = [
        {"num": 0, "titulo": "Nombre del Colegio", "icono": "🏫"},
        {"num": 1, "titulo": "Archivo 1: Nómina", "icono": "📋"},
        {"num": 2, "titulo": "Archivo 2: Notas", "icono": "📊"},
        {"num": 3, "titulo": "Descarga Final", "icono": "⬇️"}
    ]
    
    cols = st.columns(len(pasos))
    for i, paso in enumerate(pasos):
        with cols[i]:
            if paso["num"] < paso_actual:
                st.markdown(f"### ✅ {paso['icono']}")
                st.markdown(f"**{paso['titulo']}**")
                st.markdown("*Completado*")
            elif paso["num"] == paso_actual:
                st.markdown(f"### 🔵 {paso['icono']}")
                st.markdown(f"**{paso['titulo']}**")
                st.markdown("*En progreso*")
            else:
                st.markdown(f"### ⚪ {paso['icono']}")
                st.markdown(f"<span style='color: gray;'>{paso['titulo']}</span>", unsafe_allow_html=True)
                st.markdown("*Pendiente*")
    
    st.divider()

def crear_archivo_evaluador(df_archivo1, df_archivo2_procesado):
    """
    Crea el archivo evaluador haciendo un full join entre archivo1 y archivo2
    usando IDENTIFICADOR como clave única. Retorna dos DataFrames separados
    por grado: uno para 1P-3P y otro para 4P-5S.
    
    Args:
        df_archivo1: DataFrame del archivo 1 con todos los alumnos
        df_archivo2_procesado: DataFrame del archivo 2 con notas (puede ser 1P-3P o 4P-5S)
    
    Returns:
        tuple: (df_1p3p, df_4p5s) - DataFrames separados por grado
    """
    # Mapear las columnas de archivo1 a las de archivo2
    df1_base = df_archivo1[[
        "IDENTIFICADOR", 
        "PATERNO", 
        "MATERNO", 
        "NOMBRES", 
        "GRADO", 
        "SECCIÓN"
    ]].copy()
    
    # Preparar archivo2 para el merge
    df2_merge = df_archivo2_procesado.copy()
    
    # Marcar el origen de cada registro ANTES del merge
    df2_merge['_origen'] = 'archivo2'
    df1_base['_origen'] = 'archivo1'

    # Full outer join usando IDENTIFICADOR
    df_evaluador = pd.merge(
        df2_merge,
        df1_base,
        on="IDENTIFICADOR",
        how="outer",
        suffixes=("", "_archivo1"),
        indicator=True
    )
    
    # Crear columna Observaciones basada en el origen
    def asignar_observacion(row):
        if row['_merge'] == 'both':  # Aparece en ambos archivos
            return ''
        elif row['_merge'] == 'right_only':  # Solo en archivo1
            return 'SN'
        else:  # 'left_only' - Solo en archivo2
            return 'RET'
    
    df_evaluador['OBSERVACIONES'] = df_evaluador.apply(asignar_observacion, axis=1)
    
    # Eliminar columnas auxiliares
    df_evaluador = df_evaluador.drop(columns=['_merge', '_origen'], errors='ignore')
    
    # Completar datos faltantes: si no hay datos de archivo2, usar los de archivo1
    columnas_comunes = ["PATERNO", "MATERNO", "NOMBRES", "GRADO", "SECCIÓN"]
    
    # Completar primero la columna GRADO antes del filtro
    if "GRADO_archivo1" in df_evaluador.columns:
        df_evaluador["GRADO"] = df_evaluador["GRADO"].fillna(df_evaluador["GRADO_archivo1"])
        mask_vacio = (df_evaluador["GRADO"] == "") | (df_evaluador["GRADO"].isna())
        df_evaluador.loc[mask_vacio, "GRADO"] = df_evaluador.loc[mask_vacio, "GRADO_archivo1"]
    
    # Eliminar la columna temporal de GRADO_archivo1 si existe
    df_evaluador = df_evaluador.drop(columns=["GRADO_archivo1"], errors='ignore')
    
    # Continuar completando el resto de columnas comunes (excepto GRADO que ya se procesó)
    columnas_comunes_restantes = ["PATERNO", "MATERNO", "NOMBRES", "SECCIÓN"]

    for col in columnas_comunes_restantes:
        col_archivo1 = f"{col}_archivo1"
        if col_archivo1 in df_evaluador.columns:
            df_evaluador[col] = df_evaluador[col].fillna(df_evaluador[col_archivo1])
            mask_vacio = (df_evaluador[col] == "") | (df_evaluador[col].isna())
            df_evaluador.loc[mask_vacio, col] = df_evaluador.loc[mask_vacio, col_archivo1]
            df_evaluador = df_evaluador.drop(columns=[col_archivo1])
    
    # Asegurar que CURSO y NOTA VIGESIMAL existan
    if "CURSO" not in df_evaluador.columns:
        df_evaluador["CURSO"] = ""
    if "NOTA VIGESIMAL" not in df_evaluador.columns:
        df_evaluador["NOTA VIGESIMAL"] = ""
    
    # Rellenar NaN restantes con cadenas vacías
    df_evaluador = df_evaluador.fillna("")
    
    # SEPARAR EN DOS DATAFRAMES SEGÚN GRADO
    grados_1p3p = ["1P", "2P", "3P"]
    grados_4p5s = ["4P", "5P", "6P", "1S", "2S", "3S", "4S", "5S"]
    
    df_1p3p = df_evaluador[df_evaluador["GRADO"].isin(grados_1p3p)].copy()
    df_4p5s = df_evaluador[df_evaluador["GRADO"].isin(grados_4p5s)].copy()
    
    # Definir columnas finales para 1P-3P (sin NOTAS VIGESIMALES 75% ni PROMEDIO)
    columnas_1p3p = [
        "NRO.", "PATERNO", "MATERNO", "NOMBRES", "CURSO", 
        "GRADO", "SECCIÓN", "NOTA VIGESIMAL", "IDENTIFICADOR", "OBSERVACIONES"
    ]
    
    # Definir columnas finales para 4P-5S (con NOTAS VIGESIMALES 75% y PROMEDIO)
    columnas_4p5s = [
        "NRO.", "PATERNO", "MATERNO", "NOMBRES", "CURSO", 
        "GRADO", "SECCIÓN", "NOTA VIGESIMAL", 
        "NOTAS VIGESIMALES 75%", "PROMEDIO", "IDENTIFICADOR", "OBSERVACIONES"
    ]
    
    # Asegurar columnas para 1P-3P
    for col in columnas_1p3p:
        if col not in df_1p3p.columns:
            df_1p3p[col] = ""
    df_1p3p = df_1p3p[columnas_1p3p]
    
    # Asegurar columnas para 4P-5S
    for col in columnas_4p5s:
        if col not in df_4p5s.columns:
            df_4p5s[col] = ""
    df_4p5s = df_4p5s[columnas_4p5s]
    
    # Regenerar NRO. secuencial para cada DataFrame
    if len(df_1p3p) > 0:
        df_1p3p["NRO."] = range(1, len(df_1p3p) + 1)
    
    if len(df_4p5s) > 0:
        df_4p5s["NRO."] = range(1, len(df_4p5s) + 1)
    
    return df_1p3p, df_4p5s

def guardar_con_formato_original(df_procesado, archivo_original_bytes, nombre_hoja, fila_cabecera, agregar_columnas_nuevas=False, solo_hoja_especificada=False):
    """
    Preserva el formato del archivo original y actualiza solo los datos procesados
    
    Args:
        df_procesado: DataFrame con los datos procesados
        archivo_original_bytes: Bytes del archivo Excel original
        nombre_hoja: Nombre de la hoja a actualizar (None para usar la primera hoja)
        fila_cabecera: Índice de la fila donde está la cabecera (base 0 de pandas)
        agregar_columnas_nuevas: Si True, agrega columnas nuevas del df_procesado a la cabecera
    
    Returns:
        BytesIO con el archivo actualizado preservando formato
    """
    wb = load_workbook(BytesIO(archivo_original_bytes))
    
    # Si no se especifica nombre de hoja, usar la primera
    if nombre_hoja is None or nombre_hoja not in wb.sheetnames:
        ws = wb.active
    else:
        ws = wb[nombre_hoja]

    # Si solo_hoja_especificada=True, eliminar todas las demás hojas
    if solo_hoja_especificada:
      hoja_a_mantener = ws.title
      hojas_a_eliminar = [sheet for sheet in wb.sheetnames if sheet != hoja_a_mantener]
      for hoja in hojas_a_eliminar:
          wb.remove(wb[hoja])
    
    # Convertir fila_cabecera de pandas (base 0) a openpyxl (base 1)
    fila_cabecera_excel = fila_cabecera + 1
    fila_inicio_datos = fila_cabecera_excel + 1
    
    # Si se debe agregar columnas nuevas, actualizar la cabecera
    if agregar_columnas_nuevas:
        
        # Leer cabecera actual del Excel (solo celdas con valores)
        cabecera_actual = []
        ultima_col_con_datos = 0
        for idx, cell in enumerate(ws[fila_cabecera_excel], start=1):
            if cell.value is not None:
                cabecera_actual.append(str(cell.value).upper().strip())
                ultima_col_con_datos = idx
        
        cabecera_df = [str(col).upper().strip() for col in df_procesado.columns]
        
        # Encontrar columnas nuevas que no están en la cabecera actual
        columnas_nuevas = [col for col in cabecera_df if col not in cabecera_actual]
        
        # Agregar las columnas nuevas inmediatamente después de la última columna con datos
        if columnas_nuevas:
            # Obtener el estilo de la última celda de la cabecera con datos
            celda_referencia = ws.cell(row=fila_cabecera_excel, column=ultima_col_con_datos)
            
            for idx, nueva_col in enumerate(columnas_nuevas, start=1):
                nueva_celda = ws.cell(row=fila_cabecera_excel, column=ultima_col_con_datos + idx)
                nueva_celda.value = nueva_col
                
                # Copiar el estilo de la celda de referencia
                if celda_referencia.fill:
                    nueva_celda.fill = PatternFill(
                        start_color=celda_referencia.fill.start_color,
                        end_color=celda_referencia.fill.end_color,
                        fill_type=celda_referencia.fill.fill_type
                    )
                if celda_referencia.font:
                    nueva_celda.font = Font(
                        name=celda_referencia.font.name,
                        size=celda_referencia.font.size,
                        bold=celda_referencia.font.bold,
                        italic=celda_referencia.font.italic,
                        color=celda_referencia.font.color
                    )
                if celda_referencia.alignment:
                    nueva_celda.alignment = Alignment(
                        horizontal=celda_referencia.alignment.horizontal,
                        vertical=celda_referencia.alignment.vertical
                    )

    # Eliminar filas de datos antiguos (preservando cabecera y filas previas)
    if ws.max_row >= fila_inicio_datos:
        ws.delete_rows(fila_inicio_datos, ws.max_row - fila_inicio_datos + 1)
    
    # Insertar nuevos datos
    for r_idx, row in enumerate(dataframe_to_rows(df_procesado, index=False, header=False), start=fila_inicio_datos):
        for c_idx, value in enumerate(row, start=1):
            # Manejar valores NaN
            if pd.isna(value):
                value = None
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def guardar_evaluador_con_multiples_hojas(archivo_original_bytes, dict_hojas_procesadas):
    """
    Guarda un archivo Excel con múltiples hojas preservando el formato original.
    
    Args:
        archivo_original_bytes: Bytes del archivo Excel original
        dict_hojas_procesadas: Diccionario con formato {
            'nombre_hoja': {
                'df': DataFrame procesado,
                'fila_cabecera': int (índice de cabecera en base 0)
            }
        }
    
    Returns:
        BytesIO con el archivo actualizado preservando formato
    """
    wb = load_workbook(BytesIO(archivo_original_bytes))
    
    for nombre_hoja, datos in dict_hojas_procesadas.items():
        df_procesado = datos['df']
        fila_cabecera = datos['fila_cabecera']
        
        # Si la hoja no existe en el workbook, usar la primera disponible o crearla
        if nombre_hoja not in wb.sheetnames:
            # Si es la primera hoja a procesar y no existe, usar la hoja activa
            if len([k for k in dict_hojas_procesadas.keys()]) == 1:
                ws = wb.active
                ws.title = nombre_hoja
            else:
                # Crear nueva hoja
                ws = wb.create_sheet(title=nombre_hoja)
                fila_cabecera = 0  # Para hojas nuevas, empezar desde fila 0
        else:
            ws = wb[nombre_hoja]
        
        # Convertir fila_cabecera de pandas (base 0) a openpyxl (base 1)
        fila_cabecera_excel = fila_cabecera + 1
        fila_inicio_datos = fila_cabecera_excel + 1
        
        # Actualizar cabecera con las columnas del DataFrame (incluyendo OBSERVACIONES)
        cabecera_actual = []
        ultima_col_con_datos = 0
        for idx, cell in enumerate(ws[fila_cabecera_excel], start=1):
            if cell.value is not None:
                cabecera_actual.append(str(cell.value).upper().strip())
                ultima_col_con_datos = idx
        
        cabecera_df = [str(col).upper().strip() for col in df_procesado.columns]
        
        # Encontrar columnas nuevas que no están en la cabecera actual
        columnas_nuevas = [col for col in cabecera_df if col not in cabecera_actual]
        
        # Agregar las columnas nuevas
        if columnas_nuevas:
            celda_referencia = ws.cell(row=fila_cabecera_excel, column=max(1, ultima_col_con_datos))
            
            for idx, nueva_col in enumerate(columnas_nuevas, start=1):
                nueva_celda = ws.cell(row=fila_cabecera_excel, column=ultima_col_con_datos + idx)
                nueva_celda.value = nueva_col
                
                # Copiar el estilo de la celda de referencia
                if celda_referencia.fill:
                    nueva_celda.fill = PatternFill(
                        start_color=celda_referencia.fill.start_color,
                        end_color=celda_referencia.fill.end_color,
                        fill_type=celda_referencia.fill.fill_type
                    )
                if celda_referencia.font:
                    nueva_celda.font = Font(
                        name=celda_referencia.font.name,
                        size=celda_referencia.font.size,
                        bold=celda_referencia.font.bold,
                        italic=celda_referencia.font.italic,
                        color=celda_referencia.font.color
                    )
                if celda_referencia.alignment:
                    nueva_celda.alignment = Alignment(
                        horizontal=celda_referencia.alignment.horizontal,
                        vertical=celda_referencia.alignment.vertical
                    )
        
        # Eliminar filas de datos antiguos
        if ws.max_row >= fila_inicio_datos:
            ws.delete_rows(fila_inicio_datos, ws.max_row - fila_inicio_datos + 1)
        
        # Insertar nuevos datos
        for r_idx, row in enumerate(dataframe_to_rows(df_procesado, index=False, header=False), start=fila_inicio_datos):
            for c_idx, value in enumerate(row, start=1):
                if pd.isna(value):
                    value = None
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Funciones para Tab de Evaluadores:
def leer_archivo_evaluador(archivo_bytes, nombre_hoja=None):
    """Lee un archivo evaluador Excel y retorna DataFrame y metadatos"""
    try:
        wb = load_workbook(BytesIO(archivo_bytes), data_only=True)
        
        # Si no se especifica hoja, usar la primera
        if nombre_hoja is None:
            nombre_hoja = wb.sheetnames[0]
        
        if nombre_hoja not in wb.sheetnames:
            return None, f"La hoja '{nombre_hoja}' no existe en el archivo", None, None
        
        ws = wb[nombre_hoja]
        
        # Convertir a DataFrame
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Detectar cabecera usando la función existente
        fila_cabecera = detectar_cabecera_automatica(df, COLUMNAS_EVALUADOR)
        
        if fila_cabecera is None:
            return None, "No se pudo detectar la cabecera automáticamente", None, None
        
        # Extraer nombres de columnas de la fila de cabecera
        nombres_columnas_raw = df.iloc[fila_cabecera].tolist()
        
        # Limpiar nombres de columnas y manejar duplicados/None
        nombres_columnas = []
        for i, col in enumerate(nombres_columnas_raw):
            if col is None or pd.isna(col) or str(col).strip() == '' or str(col).lower() == 'nan':
                # Si la columna es None o vacía, usar un nombre genérico
                nombres_columnas.append(f"Columna_Extra_{i}")
            else:
                # Limpiar espacios extras al inicio y final
                nombres_columnas.append(str(col).strip())
        
        # Asignar columnas limpias
        df.columns = nombres_columnas
        df = df.iloc[fila_cabecera + 1:].reset_index(drop=True)
        
        # Eliminar columnas extras (las que no están en COLUMNAS_EVALUADOR)
        columnas_a_mantener = [col for col in df.columns if col in COLUMNAS_EVALUADOR]
        
        # Verificar que tengamos todas las columnas requeridas
        columnas_faltantes = [col for col in COLUMNAS_EVALUADOR if col not in columnas_a_mantener]
        if columnas_faltantes:
            return None, f"No se encontraron las columnas: {', '.join(columnas_faltantes)}. Revisa que los nombres coincidan exactamente.", None, None
        
        df = df[columnas_a_mantener]
        
        # Limpiar filas vacías
        df = df.dropna(how='all')
        
        return df, None, fila_cabecera, wb.sheetnames
        
    except Exception as e:
        return None, f"Error al leer archivo: {str(e)}", None, None

def comparar_evaluadores(df_base, df_revisar):
    """
    Compara dos archivos evaluadores.
    - Ambos deben tener las mismas columnas en el mismo orden
    - Todo debe ser idéntico EXCEPTO la columna "NOTAS VIGESIMALES 75%"
    - En el archivo BASE: pueden estar vacías "Nota Vigesimal", "NOTAS VIGESIMALES 75%", "PROMEDIO", "OBSERVACIONES"
    - En el archivo A REVISAR: "Nota Vigesimal" y "NOTAS VIGESIMALES 75%" deben estar completas
    """
    errores = []
    
    # Normalizar nombres de columnas (eliminar espacios extras pero mantener capitalización)
    df_base.columns = [str(col).strip() for col in df_base.columns]
    df_revisar.columns = [str(col).strip() for col in df_revisar.columns]
    
    columnas_base = list(df_base.columns)
    columnas_revisar = list(df_revisar.columns)
    
    # 1. Verificar que ambos tienen las columnas requeridas
    columnas_faltantes_base = [col for col in COLUMNAS_EVALUADOR if col not in columnas_base]
    columnas_faltantes_revisar = [col for col in COLUMNAS_EVALUADOR if col not in columnas_revisar]
    
    if columnas_faltantes_base:
        errores.append({
            "tipo": "error_estructura",
            "categoria": "ESTRUCTURA",
            "descripcion": f"Archivo BASE: Faltan columnas requeridas: {', '.join(columnas_faltantes_base)}",
            "archivo": "BASE",
            "fila": None,
            "columna": None,
            "valor_base": None,
            "valor_revisar": None,
            "detalle": f"Columnas actuales: {columnas_base}"
        })
    
    if columnas_faltantes_revisar:
        errores.append({
            "tipo": "error_estructura",
            "categoria": "ESTRUCTURA",
            "descripcion": f"Archivo A REVISAR: Faltan columnas requeridas: {', '.join(columnas_faltantes_revisar)}",
            "archivo": "A REVISAR",
            "fila": None,
            "columna": None,
            "valor_base": None,
            "valor_revisar": None,
            "detalle": f"Columnas actuales: {columnas_revisar}"
        })
    
    # Si faltan columnas, retornar ahora
    if columnas_faltantes_base or columnas_faltantes_revisar:
        return errores
    
    # 2. Verificar que las columnas coincidan exactamente en orden
    if columnas_base != columnas_revisar:
        errores.append({
            "tipo": "error_estructura",
            "categoria": "ESTRUCTURA",
            "descripcion": "Las columnas no coinciden entre archivos (orden diferente)",
            "archivo": "Ambos",
            "fila": None,
            "columna": None,
            "valor_base": None,
            "valor_revisar": None,
            "detalle": f"BASE: {columnas_base}\nREVISAR: {columnas_revisar}"
        })
        return errores
    
    # 3. Verificar que el número de filas sea el mismo
    if len(df_base) != len(df_revisar):
        errores.append({
            "tipo": "error_estructura",
            "categoria": "ESTRUCTURA",
            "descripcion": f"Diferente número de filas",
            "archivo": "Ambos",
            "fila": None,
            "columna": None,
            "valor_base": f"{len(df_base)} filas",
            "valor_revisar": f"{len(df_revisar)} filas",
            "detalle": None
        })
    
    # 4. Comparar todas las columnas EXCEPTO "NOTAS VIGESIMALES 75%"
    columnas_comparar = [col for col in columnas_base if col not in ["NOTAS VIGESIMALES 75%"]]
    
    for col in columnas_comparar:
        # Para estas columnas opcionales, no comparar si están vacías en BASE
        columnas_opcionales_base = ["Nota Vigesimal", "PROMEDIO", "OBSERVACIONES"]
        
        for idx in range(min(len(df_base), len(df_revisar))):
            val_base = str(df_base.loc[idx, col]).strip().upper()
            val_revisar = str(df_revisar.loc[idx, col]).strip().upper()
            
            # Normalizar valores vacíos
            valores_vacios = ["", "NAN", "NONE", "NP"]
            if val_base in valores_vacios:
                val_base = ""
            if val_revisar in valores_vacios:
                val_revisar = ""
            
            # Si la columna es opcional en BASE y está vacía en BASE, no comparar
            if col in columnas_opcionales_base and val_base == "":
                continue
            
            if val_base != val_revisar:
                # Obtener nombres y apellidos para el reporte
                paterno = str(df_base.loc[idx, "Paterno"]).strip() if "Paterno" in df_base.columns else ""
                materno = str(df_base.loc[idx, "Materno"]).strip() if "Materno" in df_base.columns else ""
                nombres = str(df_base.loc[idx, "Nombres"]).strip() if "Nombres" in df_base.columns else ""
                
                errores.append({
                    "tipo": "diferencia_contenido",
                    "categoria": "CONTENIDO DIFERENTE",
                    "paterno": paterno,
                    "materno": materno,
                    "nombres": nombres,
                    "descripcion": f"Valor diferente en columna '{col}'",
                    "archivo": "Ambos",
                    "fila": idx + 2,  # +2 por cabecera y porque index empieza en 0
                    "columna": col,
                    "valor_base": str(df_base.loc[idx, col]) if str(df_base.loc[idx, col]).strip() not in ["", "nan", "None"] else "(vacío)",
                    "valor_revisar": str(df_revisar.loc[idx, col]) if str(df_revisar.loc[idx, col]).strip() not in ["", "nan", "None"] else "(vacío)",
                    "detalle": None
                })
    
    # 5. Verificar columna "NOTAS VIGESIMALES 75%" en archivo A REVISAR
    if "NOTAS VIGESIMALES 75%" in columnas_revisar:
        for idx in range(len(df_revisar)):
            val_revisar = str(df_revisar.loc[idx, "NOTAS VIGESIMALES 75%"]).strip().upper()
            
            if val_revisar in ["", "NAN", "NONE", "NP"]:
                # Obtener nombres y apellidos para el reporte
                paterno = str(df_revisar.loc[idx, "Paterno"]).strip() if "Paterno" in df_revisar.columns else ""
                materno = str(df_revisar.loc[idx, "Materno"]).strip() if "Materno" in df_revisar.columns else ""
                nombres = str(df_revisar.loc[idx, "Nombres"]).strip() if "Nombres" in df_revisar.columns else ""
                
                errores.append({
                    "tipo": "campo_vacio_revisar",
                    "categoria": "CAMPO VACÍO EN REVISAR",
                    "paterno": paterno,
                    "materno": materno,
                    "nombres": nombres,
                    "descripcion": "Campo 'NOTAS VIGESIMALES 75%' vacío o con 'NP'",
                    "archivo": "A REVISAR",
                    "fila": idx + 2,
                    "columna": "NOTAS VIGESIMALES 75%",
                    "valor_base": str(df_base.loc[idx, "NOTAS VIGESIMALES 75%"]) if idx < len(df_base) else "N/A",
                    "valor_revisar": "(vacío)",
                    "detalle": None
                })
    
    # 6. Verificar que "Nota Vigesimal" esté completa en archivo A REVISAR
    if "Nota Vigesimal" in columnas_revisar:
        for idx in range(len(df_revisar)):
            val_revisar = str(df_revisar.loc[idx, "Nota Vigesimal"]).strip().upper()
            
            if val_revisar in ["", "NAN", "NONE", "NP"]:
                # Obtener nombres y apellidos para el reporte
                paterno = str(df_revisar.loc[idx, "Paterno"]).strip() if "Paterno" in df_revisar.columns else ""
                materno = str(df_revisar.loc[idx, "Materno"]).strip() if "Materno" in df_revisar.columns else ""
                nombres = str(df_revisar.loc[idx, "Nombres"]).strip() if "Nombres" in df_revisar.columns else ""
                
                errores.append({
                    "tipo": "campo_vacio_revisar",
                    "categoria": "CAMPO VACÍO EN REVISAR",
                    "paterno": paterno,
                    "materno": materno,
                    "nombres": nombres,
                    "descripcion": "Campo 'Nota Vigesimal' vacío o con 'NP'",
                    "archivo": "A REVISAR",
                    "fila": idx + 2,
                    "columna": "Nota Vigesimal",
                    "valor_base": str(df_base.loc[idx, "Nota Vigesimal"]) if idx < len(df_base) else "N/A",
                    "valor_revisar": "(vacío)",
                    "detalle": None
                })
    
    return errores

# ================================================
# INTERFAZ PRINCIPAL CON TABS
# ================================================

st.title("📊 Sistema de Validación de Archivos")

# Crear tabs principales
tab1, tab2 = st.tabs(["🔍 Validador General", "⚖️ Comparador de Evaluadores"])

# ================================================
# TAB 1: VALIDADOR GENERAL
# ================================================
with tab1:
    # ================================================
    # INTERFAZ PRINCIPAL
    # ================================================
    st.title("🎯 Validador de Archivos Escolares")
    st.markdown("### Sistema de Homologación de Datos")

    # Mostrar stepper
    mostrar_stepper(st.session_state.paso_actual)

    # ================================================
    # PASO 0: Nombre DEL COLEGIO
    # ================================================
    if st.session_state.paso_actual == 0:
        st.header("🏫 Paso 1: Información del Colegio")

        st.markdown("""
            <div style='background-color: #78808C; padding: 20px; border-radius: 10px; margin-bottom: 20px;'>
                <h4>Bienvenido al sistema de validación</h4>
                <p>Para comenzar, ingresa el Nombre del colegio. Este Nombre se usará para identificar los archivos descargables.</p>
            </div>
            """, unsafe_allow_html=True)

        col1, col2 = st.columns([3, 1])
        
        with col1:
            index_seleccionado = None

            if st.session_state.nombre_colegio in LISTA_COLEGIOS:
                index_seleccionado = LISTA_COLEGIOS.index(st.session_state.nombre_colegio)

            NOMBRES = st.selectbox(
                "Selecciona el colegio:",
                options=LISTA_COLEGIOS,
                index=index_seleccionado,
                placeholder="Elige un colegio..."
            )
            
        with col2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("➡️ Continuar", type="primary", use_container_width=True, disabled = not NOMBRES):
                if NOMBRES:
                    st.session_state.nombre_colegio = NOMBRES.strip()
                    st.session_state.paso_actual = 1
                    st.rerun()
                else:
                    st.error("Por favor, ingresa el Nombre del colegio")

    # ================================================
    # PASO 1: ARCHIVO 1 (NÓMINA)
    # ================================================
    elif st.session_state.paso_actual == 1:
        # Mostrar resumen del paso anterior
        with st.expander("✅ Paso 1 completado: Nombre del Colegio", expanded=False):
            st.info(f"**Colegio:** {st.session_state.nombre_colegio}")
            if st.button("🔄 Cambiar Nombre", key="cambiar_nombre"):
                st.session_state.paso_actual = 0
                st.rerun()
        
        st.header("📋 Paso 2: Archivo de Nómina de Alumnos")
        
        with st.container():
            st.markdown("""
            <div style='background-color: #78808C; padding: 20px; border-radius: 10px; margin-bottom: 20px;'>
                <h4>📄 Instrucciones</h4>
                <p>Sube el archivo Excel que contiene la nómina de alumnos.</p>
                <p><strong>Columnas requeridas:</strong></p>
                <code>Nro., Paterno, Materno, Nombres, Nacimiento (DD/MM/YYYY), Sexo (M/F), Grado, Sección, Correo institucional, Neurodiversidad (Sí/No), DNI</code>
            </div>
            """, unsafe_allow_html=True)
            
            archivo = st.file_uploader(
                "Selecciona el archivo Excel",
                type=["xls", "xlsx"],
                help="El sistema detectará automáticamente la fila de cabecera"
            )
            
            if archivo is not None:
                # GUARDAR BYTES ORIGINALES
                st.session_state.archivo1_bytes = archivo.getvalue()
                with st.spinner("🔍 Analizando archivo..."):
                    try:
                        df_original = pd.read_excel(archivo, header=None)
                        fila_detectada = detectar_cabecera_automatica(df_original, COLUMNAS_ARCHIVO1)
                        
                        if fila_detectada is not None:
                            # GUARDAR ÍNDICE DE CABECERA
                            st.session_state.archivo1_fila_cabecera = fila_detectada
                            st.success(f"✅ Cabecera detectada automáticamente en la fila {fila_detectada + 1}")
                            
                            df = pd.read_excel(archivo, header=fila_detectada)
                            
                            # Procesar columnas
                            columnas_norm = {c.strip().lower(): c for c in df.columns}
                            cols_a_usar = []
                            for col_req in COLUMNAS_ARCHIVO1:
                                col_norm = col_req.strip().lower()
                                if col_norm in columnas_norm:
                                    cols_a_usar.append(columnas_norm[col_norm])
                            
                            df = df[cols_a_usar]
                            df.columns = [col.upper() for col in COLUMNAS_ARCHIVO1]

                            # Eliminar filas con campos vacíos en PATERNO, MATERNO y NOMBRES
                            df = limpiar_filas_vacias(df, columnas_clave=["PATERNO", "MATERNO", "NOMBRES"])
                            
                            df = convertir_numericas_a_entero(df, columnas=["GRADO"])

                            # Convertir numéricas a enteros
                            df = homologar_dataframe(df)

                            # Validar campos vacíos en PATERNO, MATERNO o NOMBRES
                            columnas_obligatorias = ["PATERNO", "MATERNO", "NOMBRES"]
                            filas_vacias = df[df[columnas_obligatorias].isnull().any(axis=1)]

                            if not filas_vacias.empty:
                                st.error("❌ Se detectaron campos vacíos en nombres o apellidos (Archivo 1 - Nómina)")
                                st.dataframe(filas_vacias, use_container_width=True)
                                st.stop()
                            
                            # Validaciones para Archivo 1 (nómina)
                            errores_fatales = []
                            alertas = []
                            
                            # Validar y mapear grados
                            df, errores_grados = validar_y_mapear_grados(df, "GRADO")
                            errores_fatales.extend(errores_grados)
                            
                            # Validar sexo
                            errores_sexo = validar_sexo(df, "SEXO (M/F)")
                            alertas.extend(errores_sexo)
                            
                            # Validar secciones
                            errores_secciones = validar_secciones(df, "SECCIÓN")
                            errores_fatales.extend(errores_secciones)

                            # Validar neurodiversidad
                            errores_neuro = validar_neurodiversidad(df, "NEURODIVERSIDAD (SÍ/NO)")
                            alertas.extend(errores_neuro)
                            
                            # Validar fecha
                            errores_fecha = validar_fecha_nacimiento(df, "NACIMIENTO (DD/MM/YYYY)")
                            alertas.extend(errores_fecha)
                            
                            # Validar DNI
                            errores_dni = validar_dni(df, "DNI")
                            alertas.extend(errores_dni)
                            
                            # Validar correo
                            errores_correo = validar_correo(df, "CORREO INSTITUCIONAL")
                            alertas.extend(errores_correo)
                            
                            # Mostrar errores si existen
                            if errores_fatales:
                                st.error("❌ Se encontraron errores de validación:")
                                # Convertir lista de alertas a DataFrame
                                df_errores_fatales = pd.DataFrame(errores_fatales, columns=["Detalle de los errores críticos"])
                                    
                                # Mostrar tabla scrolleable
                                st.dataframe(
                                    df_errores_fatales,
                                    use_container_width=True,
                                    height=220  # ajusta la altura visible (unas 5-6 filas aprox)
                                )
                                    
                                st.caption(f"🔎 Total de errores: {len(errores_fatales)}")
                                st.info("Por favor, corrige estos errores en el archivo y vuelve a cargarlo")
                                st.stop()
                                
                            else:
                                df["IDENTIFICADOR"] = crear_identificador(df, "PATERNO", "MATERNO", "NOMBRES")
                                st.session_state.archivo1_df = df
                                
                                if alertas:
                                    st.warning("⚠️ Se detectaron advertencias en los datos (no bloquean el proceso):")
                                    with st.expander("Ver alertas detalladas", expanded=True):
                                        # Convertir lista de alertas a DataFrame
                                        df_alertas = pd.DataFrame(alertas, columns=["Detalle de la Alerta"])
                                        
                                        # Mostrar tabla scrolleable
                                        st.dataframe(
                                            df_alertas,
                                            use_container_width=True,
                                            height=220  # ajusta la altura visible (unas 5-6 filas aprox)
                                        )
                                        
                                        st.caption(f"🔎 Total de alertas: {len(alertas)}")
                                else:
                                    st.success("✅ Todas las validaciones pasaron correctamente")

                            if not errores_fatales:
                                # Mostrar preview
                                st.markdown("### 📊 Vista Previa de Datos")
                                st.info(f"Total de registros: {len(df)}")
                                st.dataframe(df.head(10), use_container_width=True, hide_index=True)
                            
                            # Botones de acción
                            col1, col2 = st.columns(2)
                            with col1:
                                df_descarga = df.drop(columns=["IDENTIFICADOR", "Nº"], errors="ignore")
                                buffer = guardar_con_formato_original(
                                    df_procesado=df_descarga,
                                    archivo_original_bytes=st.session_state.archivo1_bytes,
                                    nombre_hoja=None,  # Usar primera hoja
                                    fila_cabecera=st.session_state.archivo1_fila_cabecera
                                )
                                st.download_button(
                                    label="💾 Descargar Archivo Homologado",
                                    data=buffer,
                                    file_name=f"{st.session_state.nombre_colegio}_nomina_RV.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                            with col2:
                                if st.button("➡️ Continuar al Paso 3", type="primary", use_container_width=True):
                                    st.session_state.paso_actual = 2
                                    st.rerun()
                        
                        else:
                            st.warning("⚠️ No se pudo detectar la cabecera automáticamente")
                            st.markdown("### 🔍 Detección Manual")
                            st.dataframe(df_original.iloc[:15, :15], use_container_width=True)
                            
                            fila_manual = st.number_input(
                                "Indica el número de fila que contiene la cabecera:",
                                min_value=1, max_value=15, step=1
                            )
                            
                            if st.button("✔️ Validar Fila Seleccionada", type="primary"):
                                fila_idx = fila_manual - 1
                                fila = df_original.iloc[fila_idx].astype(str).str.strip().str.lower().tolist()
                                columnas_req_norm = [c.lower() for c in COLUMNAS_ARCHIVO1]
                                
                                if all(col in fila for col in columnas_req_norm):
                                    st.success("✅ Cabecera válida")
                                    df = pd.read_excel(archivo, header=fila_idx)
                                    
                                    columnas_norm = {c.strip().lower(): c for c in df.columns}
                                    cols_a_usar = []
                                    for col_req in COLUMNAS_ARCHIVO1:
                                        col_norm = col_req.strip().lower()
                                        if col_norm in columnas_norm:
                                            cols_a_usar.append(columnas_norm[col_norm])
                                    
                                    df = df[cols_a_usar]
                                    df.columns = [col.upper() for col in COLUMNAS_ARCHIVO1]
                                    
                                    # Homologar datos
                                    df = homologar_dataframe(df)
                                    
                                    # Validar campos vacíos en PATERNO, MATERNO o NOMBRES
                                    columnas_obligatorias = ["PATERNO", "MATERNO", "NOMBRES"]
                                    filas_vacias = df[df[columnas_obligatorias].isnull().any(axis=1)]

                                    if not filas_vacias.empty:
                                        st.error("❌ Se detectaron campos vacíos en nombres o apellidos (Archivo 1 - Nómina)")
                                        st.dataframe(filas_vacias, use_container_width=True)
                                        st.stop()
                                    
                                    # Validaciones para Archivo 1 (nómina)
                                    errores_fatales = []
                                    alertas = []
                                    
                                    # Validar y mapear grados
                                    df, errores_grados = validar_y_mapear_grados(df, "GRADO")
                                    errores_fatales.extend(errores_grados)
                                    
                                    # Validar sexo
                                    errores_sexo = validar_sexo(df, "SEXO (M/F)")
                                    alertas.extend(errores_sexo)
                                    
                                    # Validar secciones
                                    errores_secciones = validar_secciones(df, "SECCIÓN")
                                    errores_fatales.extend(errores_secciones)

                                    # Validar neurodiversidad
                                    errores_neuro = validar_neurodiversidad(df, "NEURODIVERSIDAD (SÍ/NO)")
                                    alertas.extend(errores_neuro)
                                    
                                    # Validar fecha
                                    errores_fecha = validar_fecha_nacimiento(df, "NACIMIENTO (DD/MM/YYYY)")
                                    alertas.extend(errores_fecha)
                                    
                                    # Validar DNI
                                    errores_dni = validar_dni(df, "DNI")
                                    alertas.extend(errores_dni)
                                    
                                    # Validar correo
                                    errores_correo = validar_correo(df, "CORREO INSTITUCIONAL")
                                    alertas.extend(errores_correo)
                                    
                                    # Mostrar errores si existen
                                    if errores_fatales:
                                        st.error("❌ Se encontraron errores de validación:")
                                        # Convertir lista de alertas a DataFrame
                                        df_errores_fatales = pd.DataFrame(errores_fatales, columns=["Detalle de la Alerta"])
                                            
                                        # Mostrar tabla scrolleable
                                        st.dataframe(
                                            df_errores_fatales,
                                            use_container_width=True,
                                            height=220  # ajusta la altura visible (unas 5-6 filas aprox)
                                        )
                                            
                                        st.caption(f"🔎 Total de errores: {len(errores_fatales)}")
                                        st.info("Por favor, corrige estos errores en el archivo y vuelve a cargarlo")
                                        st.stop()

                                    else:
                                        df["IDENTIFICADOR"] = crear_identificador(df, "PATERNO", "MATERNO", "NOMBRES")
                                        st.session_state.archivo1_df = df
                                        
                                        if alertas:
                                            st.warning("⚠️ Se detectaron advertencias en los datos (no bloquean el proceso):")
                                            with st.expander("Ver alertas detalladas", expanded=True):
                                                # Convertir lista de alertas a DataFrame
                                                df_alertas = pd.DataFrame(alertas, columns=["Detalle de la Alerta"])
                                                
                                                # Mostrar tabla scrolleable
                                                st.dataframe(
                                                    df_alertas,
                                                    use_container_width=True,
                                                    height=220  # ajusta la altura visible (unas 5-6 filas aprox)
                                                )
                                                
                                                st.caption(f"🔎 Total de alertas: {len(alertas)}")
                                        else:
                                            st.success("✅ Todas las validaciones pasaron correctamente")

                                    if not errores_fatales:
                                        # Mostrar preview
                                        st.markdown("### 📊 Vista Previa de Datos")
                                        st.info(f"Total de registros: {len(df)}")
                                        st.dataframe(df.head(10), use_container_width=True, hide_index=True)
                                    
                                    # Botones de acción
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        df_descarga = df.drop(columns=["IDENTIFICADOR", "Nº"], errors="ignore")
                                        buffer = guardar_con_formato_original(
                                            df_procesado=df_descarga,
                                            archivo_original_bytes=st.session_state.archivo1_bytes,
                                            nombre_hoja=None,  # Usar primera hoja
                                            fila_cabecera=st.session_state.archivo1_fila_cabecera
                                        )
                                        st.download_button(
                                            label="💾 Descargar Archivo Homologado",
                                            data=buffer,
                                            file_name=f"{st.session_state.nombre_colegio}_nomina_RV.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            use_container_width=True
                                        )
                                    with col2:
                                        if st.button("➡️ Continuar al Paso 3", type="primary", use_container_width=True):
                                            st.session_state.paso_actual = 2
                                            st.rerun()
                                else:
                                    st.error("❌ La fila seleccionada no contiene todas las columnas requeridas")
                    
                    except Exception as e:
                        st.error(f"❌ Error al procesar el archivo: {e}")

    # ================================================
    # PASO 2: ARCHIVO 2 (NOTAS)
    # ================================================
    elif st.session_state.paso_actual == 2:
        # Mostrar resumen de pasos anteriores
        with st.expander("✅ Pasos completados", expanded=False):
            st.success(f"**Colegio:** {st.session_state.nombre_colegio}")
            st.success(f"**Archivo 1:** {len(st.session_state.archivo1_df)} registros cargados")
            if st.button("🔙 Volver al Paso 2", key="volver_paso2"):
                st.session_state.paso_actual = 1
                st.rerun()
        
        st.header("📊 Paso 3: Archivo de Notas de Cursos")
        
        # Equivalencias de cursos
        with st.expander("⚙️ Configuración de Cursos Equivalentes", expanded=False):
            st.markdown("""
            <div style='background-color: #78808C; padding: 15px; border-radius: 10px;'>
                <p>Opcionalmente, puedes cargar un archivo .txt con cursos adicionales para reconocimiento automático.</p>
            </div>
            """, unsafe_allow_html=True)
            
            archivo_txt = st.file_uploader("Archivo de equivalencias (.txt)", type=["txt"])
            if archivo_txt:
                contenido = archivo_txt.getvalue().decode("utf-8", errors="ignore")
                nuevos = [l.strip().upper() for l in contenido.splitlines() if l.strip()]
                st.session_state.cursos_equivalentes = sorted(list(set(st.session_state.cursos_equivalentes + nuevos)))
                st.success(f"✅ {len(nuevos)} cursos agregados. Total: {len(st.session_state.cursos_equivalentes)}")
        
        # Carga del archivo
        st.markdown("""
        <div style='background-color: #78808C; padding: 20px; border-radius: 10px; margin-bottom: 20px;'>
            <h4>📄 Instrucciones</h4>
            <p>Sube el archivo Excel con las notas de los cursos.</p>
            <p><strong>Columnas requeridas:</strong></p>
            <code>Nro., Paterno, Materno, Nombres, Curso, Grado, Sección,  Nota Vigesimal</code>
        </div>
        """, unsafe_allow_html=True)
        
        archivo2 = st.file_uploader("Selecciona el archivo Excel de notas", type=["xls", "xlsx"])
        
        if archivo2 is not None:
            # GUARDAR BYTES ORIGINALES
            st.session_state.archivo2_bytes = archivo2.getvalue()
            with st.spinner("🔍 Analizando archivo y hojas disponibles..."):
                try:
                    # Leer el archivo para detectar hojas
                    xls_file = pd.ExcelFile(archivo2)
                    hojas_disponibles = xls_file.sheet_names
                    
                    # Detectar qué hojas existen
                    tiene_1p3p = "1P-3P" in hojas_disponibles
                    tiene_4p5s = "4P-5S" in hojas_disponibles
                    
                    if not tiene_1p3p and not tiene_4p5s:
                        st.error("❌ El archivo no contiene ninguna de las hojas requeridas: '1P-3P' o '4P-5S'")
                        st.info(f"Hojas encontradas: {', '.join(hojas_disponibles)}")
                        st.stop()
                    
                    # Mostrar información de hojas detectadas
                    st.success(f"✅ Hojas detectadas en el archivo, Únicas Opciones ('1P-3P' o '4P-5S'):")
                    cols_info = st.columns(2)
                    with cols_info[0]:
                        if tiene_1p3p:
                            st.info("📘 **1P-3P** encontrada")
                    with cols_info[1]:
                        if tiene_4p5s:
                            st.info("📗 **4P-5S** encontrada")
                    
                    st.divider()
                    
                    # ====================================
                    # PROCESAR HOJA 1P-3P (Solo mayúsculas)
                    # ====================================
                    df_1p3p_procesado = None
                    if tiene_1p3p:
                        st.markdown("### 📘 Procesando Hoja: 1P-3P")
                        
                        df_1p3p_original = pd.read_excel(archivo2, sheet_name="1P-3P", header=None)
                        fila_detectada_1p3p = detectar_cabecera_automatica(df_1p3p_original, COLUMNAS_ARCHIVO2)
                        
                        if fila_detectada_1p3p is not None:
                            # GUARDAR ÍNDICE DE CABECERA
                            st.session_state.archivo2_1p3p_fila_cabecera = fila_detectada_1p3p
                            st.success(f"✅ Cabecera detectada en la fila {fila_detectada_1p3p + 1}")
                            
                            df_1p3p = pd.read_excel(archivo2, sheet_name="1P-3P", header=fila_detectada_1p3p)
                            
                            # Procesar columnas
                            columnas_norm = {c.strip().lower(): c for c in df_1p3p.columns}
                            cols_a_usar = []
                            for col_req in COLUMNAS_ARCHIVO2:
                                col_norm = col_req.strip().lower()
                                if col_norm in columnas_norm:
                                    cols_a_usar.append(columnas_norm[col_norm])
                            
                            df_1p3p = df_1p3p[cols_a_usar]
                            df_1p3p.columns = [col.upper() for col in COLUMNAS_ARCHIVO2]
                            
                            # Eliminar filas con campos vacíos en PATERNO, MATERNO y NOMBRES
                            df_1p3p = limpiar_filas_vacias(df_1p3p, columnas_clave=["PATERNO", "MATERNO", "NOMBRES"])

                            # Convertir numéricas a enteros
                            df_1p3p = convertir_numericas_a_entero(df_1p3p, columnas=["GRADO", "NOTA VIGESIMAL"])

                            # Homologar datos
                            df_1p3p = homologar_dataframe(df_1p3p)

                            # Validar campos vacíos en PATERNO, MATERNO o NOMBRES
                            columnas_obligatorias = ["PATERNO", "MATERNO", "NOMBRES"]
                            filas_vacias = df_1p3p[df_1p3p[columnas_obligatorias].isnull().any(axis=1)]

                            if not filas_vacias.empty:
                                st.error("❌ Se detectaron campos vacíos en nombres o apellidos (Hoja 1P-3P)")
                                st.dataframe(filas_vacias, use_container_width=True)
                                st.stop()
                            
                            # Validaciones para Archivo 2 - Hoja 1P-3P
                            errores_validacion_1p3p = []

                            # Completar valores vacíos en NOTA VIGESIMAL con "NP"
                            if "NOTA VIGESIMAL" in df_1p3p.columns:
                                df_1p3p["NOTA VIGESIMAL"] = df_1p3p["NOTA VIGESIMAL"].fillna("NP").replace("", "NP")

                            # Validar y mapear grados
                            df_1p3p, errores_grados = validar_y_mapear_grados(df_1p3p, "GRADO", tipo_validacion="1p3p")
                            errores_validacion_1p3p.extend(errores_grados)
                            
                            # Validar secciones
                            errores_secciones = validar_secciones(df_1p3p, "SECCIÓN")
                            errores_validacion_1p3p.extend(errores_secciones)
                            
                            # Mostrar errores de validación si existen
                            if errores_validacion_1p3p:
                                st.error("❌ Errores de validación en 1P-3P:")
                                df_errores_fatales_1p3p = pd.DataFrame(errores_validacion_1p3p, columns=["Detalle de los errores críticos"])
                                    
                                # Mostrar tabla scrolleable
                                st.dataframe(
                                    df_errores_fatales_1p3p,
                                    use_container_width=True,
                                    height=220  # ajusta la altura visible (unas 5-6 filas aprox)
                                )
                                    
                                st.caption(f"🔎 Total de errores: {len(errores_validacion_1p3p)}")
                                st.info("Por favor, corrige estos errores en el archivo y vuelve a cargarlo")
                                st.stop()
                            else:
                                st.success("✅ Validaciones de grados y secciones pasadas (1P-3P)")
                            
                            # Validar cursos en 1P-3P
                            cursos_invalidos_1p3p = sorted(df_1p3p.loc[~df_1p3p["CURSO"].isin(st.session_state.cursos_equivalentes), "CURSO"].unique())
                            
                            if len(cursos_invalidos_1p3p) > 0:
                                st.warning(f"⚠️ Se detectaron {len(cursos_invalidos_1p3p)} cursos no reconocidos en 1P-3P")
                                
                                with st.form("equivalencias_form_1p3p"):
                                    st.markdown("### 🔄 Homologación de Cursos (1P-3P)")
                                    st.info("Selecciona el curso oficial correspondiente para cada curso no reconocido:")
                                    
                                    equivalencias_1p3p = {}
                                    for curso in cursos_invalidos_1p3p:
                                        equivalencias_1p3p[curso] = st.selectbox(
                                            f"📌 **{curso}**",
                                            options=["-- Seleccionar --"] + st.session_state.cursos_equivalentes,
                                            key=f"eq_1p3p_{curso}"
                                        )
                                    
                                    submitted_1p3p = st.form_submit_button("✔️ Aplicar Equivalencias (1P-3P)", type="primary")
                                    
                                    if submitted_1p3p:
                                        if any(v == "-- Seleccionar --" for v in equivalencias_1p3p.values()):
                                            st.error("❌ Debes seleccionar un curso para todos los campos")
                                        else:
                                            # Aplicar equivalencias
                                            for curso_err, curso_ok in equivalencias_1p3p.items():
                                                df_1p3p.loc[df_1p3p["CURSO"] == curso_err, "CURSO"] = curso_ok
                                            
                                            # Agregar solo IDENTIFICADOR
                                            df_1p3p["IDENTIFICADOR"] = crear_identificador(df_1p3p, "PATERNO", "MATERNO", "NOMBRES")
                                            
                                            # Reordenar
                                            cols_orden = [c for c in df_1p3p.columns if c != "IDENTIFICADOR"]
                                            cols_orden.append("IDENTIFICADOR")
                                            df_1p3p = df_1p3p[cols_orden]
                                            
                                            # Guardar en session_state
                                            st.session_state.archivo2_1p3p_df = df_1p3p
                                            st.session_state.archivo2_1p3p_df.insert(0, 'Nro.', range(1, len(st.session_state.archivo2_1p3p_df) + 1))
                                            
                                            st.success("✅ Cursos homologados correctamente en 1P-3P")
                                            st.rerun()
                            
                            # Si no hay cursos inválidos
                            if len(cursos_invalidos_1p3p) == 0 or st.session_state.archivo2_1p3p_df is not None:
                                # Usar el DataFrame guardado si existe, sino usar el actual
                                if st.session_state.archivo2_1p3p_df is not None:
                                    df_1p3p = st.session_state.archivo2_1p3p_df
                                else:
                                    # Agregar solo IDENTIFICADOR
                                    df_1p3p["IDENTIFICADOR"] = crear_identificador(df_1p3p, "PATERNO", "MATERNO", "NOMBRES")
                                    
                                    # Reordenar
                                    cols_orden = [c for c in df_1p3p.columns if c != "IDENTIFICADOR"]
                                    cols_orden.append("IDENTIFICADOR")
                                    df_1p3p = df_1p3p[cols_orden]
                                    
                                    st.session_state.archivo2_1p3p_df = df_1p3p
                                    st.session_state.archivo2_1p3p_df.insert(0, 'Nro.', range(1, len(st.session_state.archivo2_1p3p_df) + 1))
                                
                                df_1p3p_procesado = df_1p3p
                                
                                with st.expander("Vista previa 1P-3P", expanded=False):
                                    st.info(f"Total de registros: {len(df_1p3p)}")
                                    st.dataframe(df_1p3p, use_container_width=True, hide_index=True)
                        else:
                            st.error("❌ Error de cabecera en la hoja 1P-3P")
                            st.warning("⚠️ No se pudo detectar cabecera automáticamente en 1P-3P")
                            st.info("Por favor, verifica que la hoja tenga las columnas correctas:")
                            st.code("Nro., Paterno, Materno, Nombres, Curso, Grado, Sección, Nota Vigesimal")
                            st.stop()
                    
                    # ====================================
                    # PROCESAR HOJA 4P-5S (Homologación completa)
                    # ====================================
                    df_4p5s_procesado = None
                    if tiene_4p5s:
                        st.markdown("### 📗 Procesando Hoja: 4P-5S")
                        
                        df_original2 = pd.read_excel(archivo2, sheet_name="4P-5S", header=None)
                        fila_detectada2 = detectar_cabecera_automatica(df_original2, COLUMNAS_ARCHIVO2)
                        
                        if fila_detectada2 is not None:
                            # GUARDAR ÍNDICE DE CABECERA
                            st.session_state.archivo2_4p5s_fila_cabecera = fila_detectada2
                            st.success(f"✅ Cabecera detectada en la fila {fila_detectada2 + 1}")
                            
                            df2 = pd.read_excel(archivo2, sheet_name="4P-5S", header=fila_detectada2)
                        
                            # Procesar columnas
                            columnas_norm = {c.strip().lower(): c for c in df2.columns}
                            cols_a_usar = []
                            for col_req in COLUMNAS_ARCHIVO2:
                                col_norm = col_req.strip().lower()
                                if col_norm in columnas_norm:
                                    cols_a_usar.append(columnas_norm[col_norm])
                            
                            df2 = df2[cols_a_usar]
                            df2.columns = [col.upper() for col in COLUMNAS_ARCHIVO2]

                            # Eliminar filas con campos vacíos en PATERNO, MATERNO y NOMBRES
                            df2 = limpiar_filas_vacias(df2, columnas_clave=["PATERNO", "MATERNO", "NOMBRES"])
                            
                            # Convertir numéricas a enteros
                            df2 = convertir_numericas_a_entero(df2, columnas=["GRADO", "NOTA VIGESIMAL"])

                            # Homologar datos
                            df2 = homologar_dataframe(df2)

                            # Validaciones para Archivo 2 - Hoja 4P-5S
                            errores_validacion_4p5s = []

                            # Validar y mapear grados
                            df2, errores_grados = validar_y_mapear_grados(df2, "GRADO", tipo_validacion="4p5s")
                            errores_validacion_4p5s.extend(errores_grados)

                            # Validar secciones
                            errores_secciones = validar_secciones(df2, "SECCIÓN")
                            errores_validacion_4p5s.extend(errores_secciones)

                            # Mostrar errores de validación si existen
                            if errores_validacion_4p5s:
                                st.error("❌ Errores de validación en 4P-5S:")
                                df_errores_fatales_4p5s = pd.DataFrame(errores_validacion_4p5s, columns=["Detalle de los errores críticos"])
                                        
                                # Mostrar tabla scrolleable
                                st.dataframe(
                                    df_errores_fatales_4p5s,
                                    use_container_width=True,
                                    height=220  # ajusta la altura visible (unas 5-6 filas aprox)
                                )
                                        
                                st.caption(f"🔎 Total de errores: {len(errores_validacion_4p5s)}")
                                st.info("Por favor, corrige estos errores en el archivo y vuelve a cargarlo")
                                st.stop()

                            else:
                                st.success("✅ Validaciones de grados y secciones pasadas (4P-5S)")

                            # Completar valores vacíos en NOTA VIGESIMAL con "NP"
                            if "NOTA VIGESIMAL" in df2.columns:
                                df2["NOTA VIGESIMAL"] = df2["NOTA VIGESIMAL"].fillna("NP").replace("", "NP")
                            
                            # Validar campos vacíos
                            columnas_oblig = ["PATERNO", "MATERNO", "NOMBRES", "CURSO", "GRADO", "SECCIÓN", "NOTA VIGESIMAL"]
                            filas_vacias = df2[df2[columnas_oblig].isnull().any(axis=1)]
                            
                            if not filas_vacias.empty:
                                st.error("❌ Se detectaron campos vacíos")
                                st.dataframe(filas_vacias, use_container_width=True)
                                st.stop()
                            
                            # Validar cursos
                            cursos_invalidos = sorted(df2.loc[~df2["CURSO"].isin(st.session_state.cursos_equivalentes), "CURSO"].unique())
                            
                            if len(cursos_invalidos) > 0:
                                st.warning(f"⚠️ Se detectaron {len(cursos_invalidos)} cursos no reconocidos")
                                
                                with st.form("equivalencias_form"):
                                    st.markdown("### 🔄 Homologación de Cursos")
                                    st.info("Selecciona el curso oficial correspondiente para cada curso no reconocido:")
                                    
                                    equivalencias = {}
                                    for curso in cursos_invalidos:
                                        equivalencias[curso] = st.selectbox(
                                            f"📌 **{curso}**",
                                            options=["-- Seleccionar --"] + st.session_state.cursos_equivalentes,
                                            key=f"eq_{curso}"
                                        )
                                    
                                    submitted = st.form_submit_button("✔️ Aplicar Equivalencias", type="primary")
                                    
                                    if submitted:
                                        if any(v == "-- Seleccionar --" for v in equivalencias.values()):
                                            st.error("❌ Debes seleccionar un curso para todos los campos")
                                        else:
                                            # Aplicar equivalencias
                                            for curso_err, curso_ok in equivalencias.items():
                                                df2.loc[df2["CURSO"] == curso_err, "CURSO"] = curso_ok
                                            
                                            # Guardar en session_state
                                            df2["IDENTIFICADOR"] = crear_identificador(df2, "PATERNO", "MATERNO", "NOMBRES")
                                            df2["NOTAS VIGESIMALES 75%"] = ""
                                            df2["PROMEDIO"] = ""
                                            
                                            # Reordenar columnas
                                            cols_orden = [c for c in df2.columns if c != "IDENTIFICADOR"]
                                            cols_orden.append("IDENTIFICADOR")
                                            df2 = df2[cols_orden]
                                            
                                            # Guardar en session_state
                                            st.session_state.archivo2_4p5s_df = df2
                                            st.session_state.archivo2_4p5s_df.insert(0, 'Nro.', range(1, len(st.session_state.archivo2_4p5s_df) + 1))
                                            
                                            st.success("✅ Cursos homologados correctamente")
                                            st.rerun()
                        
                            # Si no hay cursos inválidos
                            if len(cursos_invalidos) == 0 or st.session_state.archivo2_4p5s_df is not None:
                                df2["IDENTIFICADOR"] = crear_identificador(df2, "PATERNO", "MATERNO", "NOMBRES")
                                df2["NOTAS VIGESIMALES 75%"] = ""
                                df2["PROMEDIO"] = ""
                                
                                # Reordenar columnas
                                cols_orden = [c for c in df2.columns if c != "IDENTIFICADOR"]
                                cols_orden.append("IDENTIFICADOR")
                                df2 = df2[cols_orden]
                                
                                st.session_state.archivo2_4p5s_df = df2
                                st.session_state.archivo2_4p5s_df.insert(0, 'Nro.', range(1, len(st.session_state.archivo2_4p5s_df) + 1))
                                
                                df_4p5s_procesado = df2

                                with st.expander("Vista previa 4P-5S", expanded=False):
                                    st.info(f"Total de registros: {len(df2)}")
                                    st.dataframe(df2, use_container_width=True, hide_index=True)
                                    
                        else:
                            st.error("❌ Error de cabecera en la hoja 4P-5S")
                            st.warning("⚠️ No se pudo detectar cabecera automáticamente en 4P-5S")
                            st.info("Por favor, verifica que la hoja tenga las columnas correctas:")
                            st.code("Nro., Paterno, Materno, Nombres, Curso, Grado, Sección, Nota Vigesimal")
                            st.stop()

                    # ====================================
                    # SECCIÓN DE DESCARGA
                    # ====================================
                    if df_1p3p_procesado is not None or df_4p5s_procesado is not None:
                        st.divider()
                        st.markdown("### 💾 Archivos Listos para Descargar")
                        
                        # Calcular número de botones de descarga
                        num_descargas = 0
                        if df_1p3p_procesado is not None:
                            num_descargas += 1  # archivo homologado 1P-3P
                        if df_4p5s_procesado is not None:
                            num_descargas += 1  # archivo homologado 4P-5S
                        # Siempre hay 1 archivo evaluador (puede tener 1 o 2 hojas)
                        num_descargas += 1
                        
                        # Diseño dinámico
                        cols_descarga = st.columns(min(num_descargas, 3))
                        col_idx = 0
                        
                        # Descargas para 1P-3P (archivo homologado)
                        if df_1p3p_procesado is not None:
                            with cols_descarga[col_idx]:
                                df_sin_notas_1p3p = df_1p3p_procesado.drop(columns=["IDENTIFICADOR", "NRO."], errors="ignore")
                                df_sin_notas_1p3p["NOTA VIGESIMAL"] = df_sin_notas_1p3p["NOTA VIGESIMAL"].astype(str).replace('NAN', 'NP')
                                buffer_1p3p = guardar_con_formato_original(
                                    df_procesado=df_sin_notas_1p3p,
                                    archivo_original_bytes=st.session_state.archivo2_bytes,
                                    nombre_hoja="1P-3P",
                                    fila_cabecera=st.session_state.archivo2_1p3p_fila_cabecera,
                                    solo_hoja_especificada=True
                                )
                                st.download_button(
                                    label="📥 1P-3P Homologado",
                                    data=buffer_1p3p,
                                    file_name=f"{st.session_state.nombre_colegio}_1P-3P_RV.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                            col_idx += 1
                        
                        # Descargas para 4P-5S (archivo homologado)
                        if df_4p5s_procesado is not None:
                            with cols_descarga[col_idx]:
                                df_sin_notas_4p5s = df_4p5s_procesado.drop(columns=["IDENTIFICADOR", "NRO.", "NOTAS VIGESIMALES 75%", "PROMEDIO"], errors="ignore")
                                df_sin_notas_4p5s["NOTA VIGESIMAL"] = df_sin_notas_4p5s["NOTA VIGESIMAL"].astype(str).replace('NAN', 'NP')
                                buffer_4p5s = guardar_con_formato_original(
                                    df_procesado=df_sin_notas_4p5s,
                                    archivo_original_bytes=st.session_state.archivo2_bytes,
                                    nombre_hoja="4P-5S",
                                    fila_cabecera=st.session_state.archivo2_4p5s_fila_cabecera,
                                    solo_hoja_especificada=True
                                )
                                st.download_button(
                                    label="📥 4P-5S Homologado",
                                    data=buffer_4p5s,
                                    file_name=f"{st.session_state.nombre_colegio}_4P-5S_RV.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                            col_idx += 1
                        
                        # ARCHIVO EVALUADOR ÚNICO (con todas las hojas necesarias)
                        with cols_descarga[col_idx]:
                            # Combinar ambos DataFrames procesados para hacer UN SOLO merge
                            df_merge_completo = pd.concat([
                                df_1p3p_procesado if df_1p3p_procesado is not None else pd.DataFrame(),
                                df_4p5s_procesado if df_4p5s_procesado is not None else pd.DataFrame()
                            ], ignore_index=True)
                            
                            # Crear evaluadores separados por grado
                            df_eval_1p3p, df_eval_4p5s = crear_archivo_evaluador(
                                st.session_state.archivo1_df,
                                df_merge_completo
                            )
                            
                            # Preparar diccionario de hojas para el archivo evaluador
                            dict_hojas_evaluador = {}
                            
                            if len(df_eval_1p3p) > 0 and df_1p3p_procesado is not None:
                                df_eval_1p3p_final = df_eval_1p3p.drop(columns=["IDENTIFICADOR"], errors="ignore")
                                df_eval_1p3p_final["NOTA VIGESIMAL"] = df_eval_1p3p_final["NOTA VIGESIMAL"].astype(str).replace('NAN', 'NP')
                                dict_hojas_evaluador["1P-3P"] = {
                                    'df': df_eval_1p3p_final,
                                    'fila_cabecera': st.session_state.archivo2_1p3p_fila_cabecera
                                }
                            
                            if len(df_eval_4p5s) > 0 and df_4p5s_procesado is not None:
                                df_eval_4p5s_final = df_eval_4p5s.drop(columns=["IDENTIFICADOR"], errors="ignore")
                                df_eval_4p5s_final["NOTA VIGESIMAL"] = df_eval_4p5s_final["NOTA VIGESIMAL"].astype(str).replace('NAN', 'NP')
                                dict_hojas_evaluador["4P-5S"] = {
                                    'df': df_eval_4p5s_final,
                                    'fila_cabecera': st.session_state.archivo2_4p5s_fila_cabecera
                                }
                            
                            # Guardar archivo evaluador con todas las hojas
                            buffer_evaluador = guardar_evaluador_con_multiples_hojas(
                                archivo_original_bytes=st.session_state.archivo2_bytes,
                                dict_hojas_procesadas=dict_hojas_evaluador
                            )
                            
                            st.download_button(
                                label="📥 Archivo Evaluador",
                                data=buffer_evaluador,
                                file_name=f"{st.session_state.nombre_colegio}_evaluador.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        
                        # Botón de finalización
                        st.divider()
                        col1, col2, col3 = st.columns([1, 1, 2])
                        with col1:
                            if st.button("✅ Finalizar Proceso", type="primary", use_container_width=True):
                                st.session_state.paso_actual = 3
                                st.rerun()
                    
                    else:
                        st.warning("⚠️ Detección manual necesaria")
                
                except Exception as e:
                    st.error(f"❌ Error: {e}")

    # ================================================
    # PASO 3: FINALIZACIÓN
    # ================================================
    elif st.session_state.paso_actual == 3:
        st.balloons()
        
        st.markdown("""
        <div style='background-color: #78808C; padding: 30px; border-radius: 15px; text-align: center;'>
            <h1>🎉 ¡Proceso Completado!</h1>
            <p style='font-size: 18px;'>Todos los archivos han sido procesados y homologados correctamente.</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 📋 Resumen del Proceso")
            st.success(f"**Colegio:** {st.session_state.nombre_colegio}")
            st.success(f"**Archivo 1:** {len(st.session_state.archivo1_df)} registros")
            
            if st.session_state.archivo2_1p3p_df is not None:
                st.success(f"**Hoja 1P-3P:** {len(st.session_state.archivo2_1p3p_df)} registros")
            if st.session_state.archivo2_4p5s_df is not None:
                st.success(f"**Hoja 4P-5S:** {len(st.session_state.archivo2_4p5s_df)} registros")
        
        with col2:
            st.markdown("### 🔄 Acciones")
            if st.button("🆕 Procesar Nuevo Colegio", type="primary", use_container_width=True):
                # Reiniciar todo
                st.session_state.paso_actual = 0
                st.session_state.nombre_colegio = ""
                st.session_state.archivo1_df = None
                st.session_state.archivo2_df = None
                st.session_state.archivo1_bytes = None
                st.session_state.archivo2_bytes = None
                st.session_state.archivo1_fila_cabecera = None
                st.session_state.archivo2_1p3p_fila_cabecera = None
                st.session_state.archivo2_4p5s_fila_cabecera = None
                st.rerun()
            
            if st.button("🔙 Volver al Paso 3", use_container_width=True):
                st.session_state.paso_actual = 2
                st.rerun()


# ================================================
# TAB 2: COMPARADOR DE EVALUADORES
# ================================================
with tab2:
    st.markdown("### Comparación de Archivos Evaluadores")
    st.info("""
    📌 **Instrucciones:**
    - Sube el archivo **BASE** (puede tener campos vacíos en: Nota Vigesimal, NOTAS VIGESIMALES 75%, PROMEDIO, OBSERVACIONES)
    - Sube el archivo **A REVISAR** (debe tener completos: Nota Vigesimal y NOTAS VIGESIMALES 75%)
    - **La única diferencia permitida** es en la columna "NOTAS VIGESIMALES 75%"
    """)
    
    col_izq, col_der = st.columns(2)
    
    # COLUMNA IZQUIERDA: Archivo Base
    with col_izq:
        st.markdown("#### 📄 Archivo BASE")
        st.caption("Campos opcionales: Nota Vigesimal, NOTAS VIGESIMALES 75%, PROMEDIO, OBSERVACIONES")
        
        archivo_base = st.file_uploader(
            "Selecciona el archivo evaluador BASE",
            type=["xlsx"],
            key="uploader_base"
        )
        
        if archivo_base:
            archivo_base_bytes = archivo_base.read()
            archivo_base.seek(0)
            
            # Detectar hojas
            xls_base = pd.ExcelFile(archivo_base)
            hojas_base = xls_base.sheet_names
            
            hoja_base_seleccionada = st.selectbox(
                "Selecciona la hoja a usar como BASE:",
                hojas_base,
                key="selector_hoja_base"
            )
            
            if st.button("✅ Cargar Archivo BASE", key="btn_cargar_base"):
                df_base, error_base, fila_cab_base, _ = leer_archivo_evaluador(
                    archivo_base_bytes,
                    hoja_base_seleccionada
                )
                
                if error_base:
                    st.error(f"❌ {error_base}")
                else:
                    st.session_state.comparador_archivo_base = {
                        'df': df_base,
                        'nombre_hoja': hoja_base_seleccionada,
                        'fila_cabecera': fila_cab_base
                    }
                    st.success(f"✅ Archivo BASE cargado ({len(df_base)} registros)")
                    st.success(f"🔍 Cabecera detectada en fila {fila_cab_base + 1}")
            
            if st.session_state.comparador_archivo_base:
                with st.expander("👁️ Vista previa - Archivo BASE"):
                    st.dataframe(st.session_state.comparador_archivo_base['df'].head(10), hide_index=True)
    
    # COLUMNA DERECHA: Archivo a Revisar
    with col_der:
        st.markdown("#### 📝 Archivo A REVISAR")
        st.caption("Debe tener completos: Nota Vigesimal y NOTAS VIGESIMALES 75%")
        
        archivo_revisar = st.file_uploader(
            "Selecciona el archivo evaluador A REVISAR",
            type=["xlsx"],
            key="uploader_revisar"
        )
        
        if archivo_revisar:
            archivo_revisar_bytes = archivo_revisar.read()
            archivo_revisar.seek(0)
            
            # Detectar hojas
            xls_revisar = pd.ExcelFile(archivo_revisar)
            hojas_revisar = xls_revisar.sheet_names
            
            hoja_revisar_seleccionada = st.selectbox(
                "Selecciona la hoja a REVISAR:",
                hojas_revisar,
                key="selector_hoja_revisar"
            )
            
            if st.button("✅ Cargar Archivo A REVISAR", key="btn_cargar_revisar"):
                df_revisar, error_revisar, fila_cab_revisar, _ = leer_archivo_evaluador(
                    archivo_revisar_bytes,
                    hoja_revisar_seleccionada
                )
                
                if error_revisar:
                    st.error(f"❌ {error_revisar}")
                else:
                    st.session_state.comparador_archivo_revisar = {
                        'df': df_revisar,
                        'nombre_hoja': hoja_revisar_seleccionada,
                        'fila_cabecera': fila_cab_revisar
                    }
                    st.success(f"✅ Archivo A REVISAR cargado ({len(df_revisar)} registros)")
                    st.success(f"🔍 Cabecera detectada en fila {fila_cab_revisar + 1}")
            
            if st.session_state.comparador_archivo_revisar:
                with st.expander("👁️ Vista previa - Archivo A REVISAR"):
                    st.dataframe(st.session_state.comparador_archivo_revisar['df'].head(10), hide_index=True)
    
    # SECCIÓN DE COMPARACIÓN
    st.divider()
    
    if st.session_state.comparador_archivo_base and st.session_state.comparador_archivo_revisar:
        col_comp1, col_comp2, col_comp3 = st.columns([1, 2, 1])
        
        with col_comp2:
            if st.button("🔍 COMPARAR ARCHIVOS", type="primary", use_container_width=True):
                with st.spinner("Comparando archivos..."):
                    errores = comparar_evaluadores(
                        st.session_state.comparador_archivo_base['df'].copy(),
                        st.session_state.comparador_archivo_revisar['df'].copy()
                    )
                    st.session_state.comparador_resultados = errores
        
        # MOSTRAR RESULTADOS
        if st.session_state.comparador_resultados:
            st.divider()
            
            # Si NO hay errores, mostrar éxito
            if len(st.session_state.comparador_resultados) == 0:
                st.success("🎉 **¡VALIDACIÓN EXITOSA!**")
                st.success("✅ Los archivos son idénticos excepto por la columna 'NOTAS VIGESIMALES 75%', que está correctamente completada en el archivo A REVISAR.")
                st.balloons()
            else:
                # Hay errores: mostrar tabla
                st.error("❌ **SE ENCONTRARON DIFERENCIAS**")
                st.warning(f"⚠️ Total de problemas encontrados: **{len(st.session_state.comparador_resultados)}**")
                
                st.markdown("### 📋 Tabla de Errores a Corregir")
                st.info("Descarga esta tabla, corrige los errores en tu archivo y vuelve a subirlo.")
                
                # Crear DataFrame para mostrar errores
                df_errores = pd.DataFrame(st.session_state.comparador_resultados)
                
                # Reordenar columnas para mejor visualización
                columnas_orden = ['categoria', 'fila', 'paterno', 'materno', 'nombres', 'columna', 'descripcion', 'archivo', 'valor_base', 'valor_revisar']
                columnas_disponibles = [col for col in columnas_orden if col in df_errores.columns]
                df_errores_ordenado = df_errores[columnas_disponibles]
                
                # Renombrar columnas para mejor presentación
                df_errores_ordenado = df_errores_ordenado.rename(columns={
                    'categoria': 'CATEGORÍA',
                    'fila': 'FILA',
                    'paterno': 'APELLIDO PATERNO',
                    'materno': 'APELLIDO MATERNO',
                    'nombres': 'NOMBRES',
                    'columna': 'COLUMNA',
                    'descripcion': 'DESCRIPCIÓN',
                    'archivo': 'ARCHIVO',
                    'valor_base': 'VALOR BASE',
                    'valor_revisar': 'VALOR A REVISAR'
                })
                
                # Mostrar tabla
                st.dataframe(
                    df_errores_ordenado,
                    use_container_width=True,
                    #height=400
                )
                
                # Botón para descargar reporte
                st.divider()
                col_desc1, col_desc2, col_desc3 = st.columns([1, 1, 1])
                with col_desc2:
                    csv = df_errores_ordenado.to_csv(index=False).encode('utf-8-sig')
                    st.download_button(
                        label="📥 Descargar Reporte de Errores (CSV)",
                        data=csv,
                        file_name="reporte_errores_comparacion.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
    
    else:
        st.info("👆 Carga ambos archivos para comenzar la comparación")
    
    # Botón para resetear comparador
    st.divider()
    if st.button("🔄 Limpiar y Comenzar Nueva Comparación", key="btn_reset_comparador"):
        st.session_state.comparador_archivo_base = None
        st.session_state.comparador_archivo_revisar = None
        st.session_state.comparador_resultados = None
        st.rerun()